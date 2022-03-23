import openpyxl


inv_file = openpyxl.load_workbook('inventory.xlsx')

product_list = inv_file["Sheet1"]


products_per_supplier = {}
total_value_per_supplier = {}
pro_inv_under_10 = {}


# def get_products_per_supplier(product_list)
for row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(row, 4).value
    inventory = product_list.cell(row, 2).value
    price = product_list.cell(row, 3).value
    product_num = product_list.cell(row, 1).value
    inventory_price = product_list.cell(row, 5)
    # calc of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1

    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

    # calc of value per supplier
    if supplier_name in total_value_per_supplier:
        current_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_value + \
            (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # calc of products under 10
    if inventory < 10:
        pro_inv_under_10[int(product_num)] = int(inventory)

    # calc of inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)

inv_file.save('inventory_with_total.xlsx')
